#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
多 ETF 公示份额自动跟踪脚本(清单驱动)

数据来源: 按 etf清单.xlsx 中的"来源"选择对应交易所的数据接口, 模拟浏览器请求取数:
  - 上交所:  https://www.sse.com.cn/market/funddata/volumn/etfvolumn/
      接口 POST https://query.sse.com.cn/commonQuery.do
      sqlId=COMMON_SSE_ZQPZ_ETFZL_XXPL_ETFGM_SEARCH_L, STAT_DATE 留空=最新交易日
      字段: STAT_DATE(日期) / SEC_CODE / TOT_VOL(总份额,万份)
  - 深交所:  https://www.szse.cn/market/fund/volume/etf/index.html
      接口 GET https://www.szse.cn/api/report/ShowReport/data
      CATALOGID=scsj_fund_jjgm&jjlb=ETF&txtDm=代码&txtStart/txtEnd=日期
      字段: size_date(日期) / fund_code / security_short_name / current_size(基金规模,万份)
      日期需显式指定: 从"上一个工作日"起向前最多回退10天找有数据的一天(覆盖周末/节假日)

行为:
  1. 读取 etf清单.xlsx(名称|代码|来源|网址), 遍历每个基金
  2. 按来源调用对应接口, 取"最新交易日"的公示份额
  3. 打开 etf跟踪.xlsx: 每个基金一个 sheet(sheet名=代码, 不存在自动创建并写表头)
  4. 若该日期已存在则跳过(幂等); 否则追加一行并写入计算列公式:
        F 当日变化         = E(本行) - E(上一行)
        G 环比变动比例     = F(本行) / E(上一行)
        H 基期总份额       = 沿用上一行(若无则留空, I 列也不写)
        I 相对基期变动比例 = F(本行) / H(本行)

节假日处理: 任务在每个交易日早 6 点运行。若前一日为节假日/周末, 接口返回的
最新交易日就是节前最后交易日, 脚本自动追加该日数据(若 Excel 尚无); 已有则跳过。

用法: python update_etf.py [--force] | [--backfill 开始日期 结束日期]
  --force   跳过"周末非交易日"判断强制运行(用于手动补数据/测试)
  --backfill 开始日期 结束日期  历史回填模式: 重建每个基金 sheet 为指定区间
              的全部交易日数据(升序)并统一写入公式; 同时自动抓取各基金在
              BASE_DATE(默认2025-12-31)的份额作为基期总份额(H列)
"""
import sys
import os
import json
import re
import time
import logging
import random
from datetime import datetime, date, timedelta

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = r"D:\myWorkSpace\myProject\汇金跟踪"
LIST_PATH = os.environ.get("ETF_LIST_PATH", BASE_DIR + r"\etf清单.xlsx")
TRACK_PATH = os.environ.get("ETF_TRACK_PATH", BASE_DIR + r"\etf跟踪.xlsx")
LIST_SHEET = "ETF清单"

COL_HEADERS = ["来源", "代码", "名称", "日期", "当日公示份额（万份）",
               "当日变化", "环比变动比例", "基期总份额", "相对基期变动比例"]

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0 Safari/537.36")
HEADERS_SSE = {"User-Agent": UA,
               "Referer": "https://www.sse.com.cn/market/funddata/volumn/etfvolumn/",
               "Accept": "*/*"}
HEADERS_SZSE = {"User-Agent": UA,
                "Referer": "https://www.szse.cn/market/fund/volume/etf/index.html",
                "Accept": "*/*"}

SSE_API_URL = "https://query.sse.com.cn/commonQuery.do?jsonCallBack=callback"
SSE_SQLID = "COMMON_SSE_ZQPZ_ETFZL_XXPL_ETFGM_SEARCH_L"
SZSE_API_URL = "https://www.szse.cn/api/report/ShowReport/data"
SZSE_CATALOGID = "scsj_fund_jjgm"
BASE_DATE = "2025-12-31"   # 基期日: 各基金此日的公示份额作为"基期总份额"(H列)
SUMMARY_SHEET = "总结"     # 汇总表: 每基金一行, 从各明细表最新行取值

SUMMARY_HEADERS = ["代码", "名称", "最新日期", "公示份额", "当日变化",
                   "环比昨日变动", "相对基期变动", "状态", "连续n日累计变动比例"]

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("etf-track")


# ---------- 基础数据 ----------

def load_funds():
    """读取清单: 返回 [{name, code, source, url}, ...]"""
    wb = openpyxl.load_workbook(LIST_PATH)
    ws = wb[LIST_SHEET]
    funds = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None] * 4
        name, code, source, url = row[0], row[1], row[2], row[3]
        if not code or not str(code).strip():
            continue
        funds.append({
            "name": str(name).strip() if name else "",
            "code": str(code).strip(),
            "source": str(source).strip() if source else "",
            "url": str(url).strip() if url else "",
        })
    return funds


# ---------- 数据抓取 ----------

def fetch_sse(code: str) -> tuple:
    """上交所: 返回 (最新交易日, 总份额万份)。STAT_DATE 留空取最新。"""
    params = {
        "isPagination": "true",
        "pageHelp.pageSize": "100",
        "pageHelp.pageNo": "1",
        "pageHelp.beginPage": "1",
        "pageHelp.cacheSize": "1",
        "pageHelp.endPage": "1",
        "sqlId": SSE_SQLID,
        "STAT_DATE": "",
    }
    resp = requests.post(SSE_API_URL, data=params, headers=HEADERS_SSE, timeout=30)
    resp.raise_for_status()
    resp.encoding = "utf-8"
    text = resp.text.strip()
    m = re.match(r"^[^(]*\((.*)\)\s*;?\s*$", text, re.S)
    if not m:
        raise RuntimeError("上交所接口返回无法解析(JSONP): " + text[:200])
    data = json.loads(m.group(1))
    rows = data.get("result") or data.get("pageHelp", {}).get("data") or []
    if not rows:
        raise RuntimeError("上交所接口未返回数据")
    target = next((r for r in rows if str(r.get("SEC_CODE", "")).strip() == code), None)
    if not target:
        raise RuntimeError("上交所最新交易日(%s)中未找到 %s" % (rows[0].get("STAT_DATE"), code))
    return target["STAT_DATE"], float(target["TOT_VOL"])


def fetch_szse(code: str) -> tuple:
    """深交所: 返回 (交易日, 基金规模万份)。日期需显式指定, 从上一个工作日向前回退。"""
    d = date.today() - timedelta(days=1)
    while d.weekday() >= 5:          # 回到最近的工作日
        d -= timedelta(days=1)
    for _ in range(10):              # 最多回退10天, 覆盖节假日
        ds = d.strftime("%Y-%m-%d")
        params = {
            "SHOWTYPE": "JSON",
            "CATALOGID": SZSE_CATALOGID,
            "jjlb": "ETF",
            "txtStart": ds,
            "txtEnd": ds,
            "txtDm": code,
            "tab1PAGENO": "1",
            "tab1PAGESIZE": "20",
            "RANDOM": "%.10f" % random.random(),
        }
        resp = requests.get(SZSE_API_URL, params=params, headers=HEADERS_SZSE, timeout=30)
        resp.raise_for_status()
        resp.encoding = "utf-8"
        data = resp.json()
        rows = (data[0].get("data") or []) if isinstance(data, list) and data else []
        for r in rows:
            if str(r.get("fund_code", "")).strip() == code:
                return ds, float(str(r.get("current_size", "")).replace(",", ""))
        d -= timedelta(days=1)
    raise RuntimeError("深交所近10天内未找到 %s 的规模数据" % code)


def fetch_sse_history(code: str, start: str, end: str) -> dict:
    """上交所历史: 逐工作日查询 [start,end], 返回 {日期: 份额万份}。节假日无数据自动跳过。"""
    result = {}
    d = datetime.strptime(start, "%Y-%m-%d").date()
    end_d = datetime.strptime(end, "%Y-%m-%d").date()
    while d <= end_d:
        if d.weekday() < 5:                    # 仅工作日尝试(节假日接口返回空自动跳过)
            ds = d.strftime("%Y-%m-%d")
            params = {
                "isPagination": "true",
                "pageHelp.pageSize": "200",
                "pageHelp.pageNo": "1",
                "pageHelp.beginPage": "1",
                "pageHelp.cacheSize": "1",
                "pageHelp.endPage": "1",
                "sqlId": SSE_SQLID,
                "STAT_DATE": ds,
            }
            for attempt in range(3):           # 接口偶发超时, 重试最多3次
                try:
                    resp = requests.post(SSE_API_URL, data=params, headers=HEADERS_SSE, timeout=30)
                    resp.raise_for_status()
                    resp.encoding = "utf-8"
                    m = re.match(r"^[^(]*\((.*)\)\s*;?\s*$", resp.text.strip(), re.S)
                    if m:
                        data = json.loads(m.group(1))
                        rows = data.get("result") or data.get("pageHelp", {}).get("data") or []
                        for r in rows:
                            if str(r.get("SEC_CODE", "")).strip() == code:
                                result[ds] = float(r["TOT_VOL"])
                                break
                    break
                except Exception as e:
                    if attempt == 2:
                        log.warning("[%s] %s 查询失败(已重试3次): %s", code, ds, e)
                    else:
                        time.sleep(1.5)
            time.sleep(0.3)                    # 温和限速
        d += timedelta(days=1)
    return result


def fetch_szse_history(code: str, start: str, end: str) -> dict:
    """深交所历史: 日期范围查询+翻页, 返回 {日期: 份额万份}。"""
    result = {}
    page = 1
    while True:
        params = {
            "SHOWTYPE": "JSON",
            "CATALOGID": SZSE_CATALOGID,
            "jjlb": "ETF",
            "txtStart": start,
            "txtEnd": end,
            "txtDm": code,
            "tab1PAGENO": str(page),
            "tab1PAGESIZE": "100",
            "RANDOM": "%.10f" % random.random(),
        }
        resp = requests.get(SZSE_API_URL, params=params, headers=HEADERS_SZSE, timeout=30)
        resp.raise_for_status()
        resp.encoding = "utf-8"
        data = resp.json()
        meta = data[0]["metadata"]
        rows = data[0].get("data") or []
        for r in rows:
            if str(r.get("fund_code", "")).strip() == code:
                result[r["size_date"]] = float(str(r.get("current_size", "")).replace(",", ""))
        if not rows or len(result) >= int(meta.get("recordcount") or 0):
            break
        page += 1
    return result


def fetch_base_share(fund: dict):
    """查询基金在 BASE_DATE 的份额作为基期总份额(H列)。失败返回 None。"""
    try:
        if fund["source"] == "上交所":
            h = fetch_sse_history(fund["code"], BASE_DATE, BASE_DATE)
        elif fund["source"] == "深交所":
            h = fetch_szse_history(fund["code"], BASE_DATE, BASE_DATE)
        else:
            return None
        return h.get(BASE_DATE)
    except Exception:
        return None


# ---------- Excel 写入 ----------

def ensure_sheet(wb, code: str):
    """获取/创建代码对应的 sheet, 新建时写入表头。"""
    if code in wb.sheetnames:
        return wb[code]
    ws = wb.create_sheet(code)
    for c, h in enumerate(COL_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    return ws


def update_sheet(wb, fund: dict, stat_date: str, share: float) -> str:
    """追加一行到对应基金 sheet, 写入计算列公式。返回结果描述。"""
    code = fund["code"]
    ws = ensure_sheet(wb, code)

    last_row = 1                       # 表头行视为基准; 仅有表头时 last_row=1
    for r in range(ws.max_row, 0, -1):
        if ws.cell(row=r, column=1).value is not None:
            last_row = r
            break

    d = datetime.strptime(stat_date, "%Y-%m-%d")
    for r in range(2, last_row + 1):   # 幂等去重
        dv = ws.cell(row=r, column=4).value
        if isinstance(dv, (datetime, date)) and dv.date() == d.date():
            return "[%s] %s 已存在(第%d行), 跳过追加" % (code, stat_date, r)

    # B 列格式跟随该 sheet 已有行(兼容历史制表符格式)
    b_prefix = ""
    if last_row >= 2:
        bv = ws.cell(row=2, column=2).value
        if isinstance(bv, str) and bv.startswith("\t"):
            b_prefix = "\t"

    new_row = last_row + 1
    ws.cell(row=new_row, column=1, value="上交所" if fund["source"] == "上交所" else "深交所")
    ws.cell(row=new_row, column=2, value=b_prefix + code)
    ws.cell(row=new_row, column=3, value=fund["name"])
    ws.cell(row=new_row, column=4, value=d)
    ws.cell(row=new_row, column=5, value=share)

    if last_row >= 2:                  # 有前一条数据行才写公式
        ws.cell(row=new_row, column=6, value="=E%d-E%d" % (new_row, last_row))
        ws.cell(row=new_row, column=7, value="=F%d/E%d" % (new_row, last_row))
        base = ws.cell(row=last_row, column=8).value
        if base is not None:
            ws.cell(row=new_row, column=8, value=base)
            ws.cell(row=new_row, column=9, value="=F%d/H%d" % (new_row, new_row))

    return "[%s] 已追加 %s 份额 %.2f 万份 → 第%d行" % (code, stat_date, share, new_row)


def rebuild_sheet(wb, fund: dict, records: dict, base, period_desc: str) -> str:
    """重建基金 sheet: 表头 + records(日期升序全量) + 公式 + H基期。返回结果描述。"""
    code = fund["code"]
    if code in wb.sheetnames:
        del wb[code]
    ws = wb.create_sheet(code)
    for c, h in enumerate(COL_HEADERS, start=1):     # 表头
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="B50005")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sorted_dates = sorted(records.keys())
    for i, ds in enumerate(sorted_dates):
        r = i + 2
        ws.cell(row=r, column=1, value="上交所" if fund["source"] == "上交所" else "深交所")
        ws.cell(row=r, column=2, value=code)
        ws.cell(row=r, column=3, value=fund["name"])
        ws.cell(row=r, column=4, value=datetime.strptime(ds, "%Y-%m-%d"))
        ws.cell(row=r, column=5, value=records[ds])
        if i >= 1:                                   # 第二条起写计算公式
            ws.cell(row=r, column=6, value="=E%d-E%d" % (r, r - 1))
            ws.cell(row=r, column=7, value="=F%d/E%d" % (r, r - 1))
        if base is not None:                         # 基期总份额(每行) + 相对基期比例
            ws.cell(row=r, column=8, value=base)
            if i >= 1:
                ws.cell(row=r, column=9, value="=F%d/H%d" % (r, r))
    return "[%s] 重建完成: %d 个交易日 (%s)%s" % (
        code, len(sorted_dates), period_desc,
        ", 基期%.2f" % base if base is not None else ", 无基期(留空)")


# ---------- 总结表 ----------

def calc_status_from_detail(ws) -> tuple:
    """依据明细 sheet 的 E(份额)/H(基期) 自算相对基期变动序列, 判定状态。
    返回 (状态文本, n)。n 用于"连续n日累计变动比例"(反转时 n=1)。"""
    vals = []                                    # 相对基期变动序列 (r>=3)
    for r in range(3, ws.max_row + 1):
        e_prev = ws.cell(row=r - 1, column=5).value
        e = ws.cell(row=r, column=5).value
        h = ws.cell(row=r, column=8).value
        if not all(isinstance(v, (int, float)) for v in (e_prev, e, h)) or h == 0:
            continue
        vals.append((r, (e - e_prev) / h))
    if len(vals) < 2:                            # 不足两日, 无法判定四种状态
        return "-", None
    cur, prev = vals[-1][1], vals[-2][1]
    if cur >= 0:
        if prev < 0:
            return "反转增持", 1
        x = 0
        for _, v in reversed(vals):
            if v >= 0:
                x += 1
            else:
                break
        return "连续%d日增持" % x, x
    else:
        if prev >= 0:
            return "反转减持", 1
        x = 0
        for _, v in reversed(vals):
            if v < 0:
                x += 1
            else:
                break
        return "连续%d日减持" % x, x


def _sheet_ref(name: str) -> str:
    """构造公式中的 sheet 引用名(数字/特殊字符开头需加单引号)。"""
    return "'%s'" % name if not name[0].isalpha() else name


def update_summary(wb, funds: list) -> list:
    """更新"总结"sheet: 每基金一行。
    - C~G 列: 动态公式引用明细表最新行(INDEX+COUNTA), 明细新增后自动跟随
    - H 状态 / I 累计: 依据明细历史计算(脚本算好写入)
    """
    if SUMMARY_SHEET in wb.sheetnames:
        ws = wb[SUMMARY_SHEET]
    else:
        ws = wb.create_sheet(SUMMARY_SHEET)
        for c, h in enumerate(SUMMARY_HEADERS, start=1):
            ws.cell(row=1, column=c, value=h)
    if ws.max_row == 1:                          # 只有表头时补写表头样式(若新建已写, 无妨)
        for c, h in enumerate(SUMMARY_HEADERS, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="B50005")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 已有行: 代码(B列) → 行号
    row_map = {}
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a:
            row_map[str(a).strip()] = r

    used_rows = []
    for fund in funds:
        code = fund["code"]
        if code not in wb.sheetnames:            # 明细表不存在则跳过
            continue
        ref = _sheet_ref(code)
        r = row_map.get(code)
        if r is None:                            # 新基金追加行
            r = ws.max_row + 1
        used_rows.append(r)
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=fund["name"])
        ws.cell(row=r, column=3, value="=INDEX(%s!$D:$D,COUNTA(%s!$D:$D))" % (ref, ref))
        ws.cell(row=r, column=4, value="=INDEX(%s!$E:$E,COUNTA(%s!$E:$E))" % (ref, ref))
        ws.cell(row=r, column=5, value="=INDEX(%s!$F:$F,COUNTA(%s!$E:$E))" % (ref, ref))
        ws.cell(row=r, column=6, value="=INDEX(%s!$G:$G,COUNTA(%s!$E:$E))" % (ref, ref))
        ws.cell(row=r, column=7, value="=INDEX(%s!$I:$I,COUNTA(%s!$E:$E))" % (ref, ref))
        ws.cell(row=r, column=8, value="")
        status, n = calc_status_from_detail(wb[code])
        ws.cell(row=r, column=8, value=status)
        if n:
            ws.cell(row=r, column=9,
                    value="=SUM(INDEX(%s!$I:$I,COUNTA(%s!$E:$E)-%d+1):INDEX(%s!$I:$I,COUNTA(%s!$E:$E)))"
                          % (ref, ref, n, ref, ref))
        else:
            ws.cell(row=r, column=9, value=None)
        # 格式: C日期 / D,E数值 / F,G,I百分比
        ws.cell(row=r, column=3).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=4).number_format = "#,##0.00"
        ws.cell(row=r, column=5).number_format = "#,##0.00"
        ws.cell(row=r, column=6).number_format = "0.00%"
        ws.cell(row=r, column=7).number_format = "0.00%"
        ws.cell(row=r, column=9).number_format = "0.00%"
    return used_rows


def main():
    # ---- 历史回填模式 ----
    if "--backfill" in sys.argv:
        idx = sys.argv.index("--backfill")
        start, end = sys.argv[idx + 1], sys.argv[idx + 2]
        only_code = None
        if "--only" in sys.argv:
            only_code = sys.argv[sys.argv.index("--only") + 1]
        funds = load_funds()
        if only_code:
            funds = [f for f in funds if f["code"] == only_code]
        wb = openpyxl.load_workbook(TRACK_PATH)
        log.info("回填区间 %s ~ %s, 共 %d 只基金%s", start, end, len(funds),
                 " (仅 %s)" % only_code if only_code else "")
        for fund in funds:
            code, source = fund["code"], fund["source"]
            try:
                base = fetch_base_share(fund)
                log.info("[%s] 基期(%s)份额: %s", code, BASE_DATE,
                         ("%.2f 万份" % base) if base else "未获取到(将留空)")
                if source == "上交所":
                    hist = fetch_sse_history(code, start, end)
                elif source == "深交所":
                    hist = fetch_szse_history(code, start, end)
                else:
                    log.error("[%s] 未知来源 %r, 跳过", code, source)
                    continue
                log.info("[%s] 区间内取到 %d 个交易日数据", code, len(hist))
                log.info(rebuild_sheet(wb, fund, hist, base, "%s~%s" % (start, end)))
            except Exception as e:
                log.error("[%s] 回填失败: %s", code, e)
        update_summary(wb, funds)
        wb.save(TRACK_PATH)
        log.info("回填完成, 已保存: %s", TRACK_PATH)
        return
    # ---- 日常更新模式 ----
    force = "--force" in sys.argv
    today = date.today()
    log.info("运行日期: %s (%s)", today, "一二三四五六日"[today.weekday()])
    if not force and today.weekday() >= 5:
        log.info("今天是周末, 非交易日, 跳过。")
        return

    funds = load_funds()
    log.info("共 %d 只基金待更新: %s", len(funds),
             ", ".join("%s(%s)" % (f["code"], f["source"]) for f in funds))

    wb = openpyxl.load_workbook(TRACK_PATH)
    results = []
    for fund in funds:
        code, source = fund["code"], fund["source"]
        try:
            if source == "上交所":
                sd, share = fetch_sse(code)
            elif source == "深交所":
                sd, share = fetch_szse(code)
            else:
                results.append("[%s] 未知来源 %r, 跳过" % (code, source))
                continue
            log.info("[%s] %s 公示份额 %.2f 万份", code, sd, share)
            results.append(update_sheet(wb, fund, sd, share))
        except Exception as e:
            log.error("[%s] 失败: %s", code, e)
            results.append("[%s] 失败: %s" % (code, e))
    update_summary(wb, funds)
    wb.save(TRACK_PATH)

    log.info("---- 本次结果汇总 ----")
    for line in results:
        log.info(line)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log.error("任务失败: %s", e)
        sys.exit(1)
