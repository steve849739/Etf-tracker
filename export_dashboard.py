#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导出 ETF 跟踪数据为自包含 HTML 展示页。

读 etf跟踪.xlsx 的明细 sheet(每基金一个), 自算汇总与明细数据(不依赖公式缓存),
生成 dashboard/report.html —— 数据内嵌, 双击浏览器打开即用。

用法:
  python export_dashboard.py

环境变量: ETF_TRACK_PATH(跟踪表路径) ETF_OUT_HTML(输出HTML路径, 默认 dashboard/report.html)
"""
import os
import json
import logging
from datetime import datetime, date

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TRACK_PATH = os.environ.get("ETF_TRACK_PATH", os.path.join(BASE_DIR, "etf跟踪.xlsx"))
OUT_HTML = os.environ.get("ETF_OUT_HTML", os.path.join(BASE_DIR, "dashboard", "report.html"))
SUMMARY_SHEET = "总结"
SKIP_SHEETS = {"总结"}                      # 非明细 sheet 跳过

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("etf-export")


# ---------- 读取与计算 ----------

def calc_status_from_detail(ws) -> tuple:
    """同 etf_update.py: 用 E(份额)/H(基期) 自算相对基期变动序列判定状态。"""
    vals = []
    for r in range(3, ws.max_row + 1):
        ep = ws.cell(row=r - 1, column=5).value
        e = ws.cell(row=r, column=5).value
        h = ws.cell(row=r, column=8).value
        if not all(isinstance(v, (int, float)) for v in (ep, e, h)) or h == 0:
            continue
        vals.append((e - ep) / h)
    if len(vals) < 2:
        return "-", None
    cur, prev = vals[-1], vals[-2]
    if cur >= 0:
        if prev < 0:
            return "反转增持", 1
        x = 0
        for v in reversed(vals):
            if v >= 0: x += 1
            else: break
        return "连续%d日增持" % x, x
    else:
        if prev >= 0:
            return "反转减持", 1
        x = 0
        for v in reversed(vals):
            if v < 0: x += 1
            else: break
        return "连续%d日减持" % x, x


def _rel_base(ws, r):
    """相对基期变动: (E本 - E上) / H。无前一日时返回 None。"""
    ep = ws.cell(row=r - 1, column=5).value
    e = ws.cell(row=r, column=5).value
    h = ws.cell(row=r, column=8).value
    if not all(isinstance(v, (int, float)) for v in (ep, e, h)) or h == 0:
        return None
    return (e - ep) / h


def load_detail(ws) -> list:
    """读取明细 sheet 全部数据行(日期升序), 自算计算列。"""
    rows = []
    prev_e = None
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        if not isinstance(d, (datetime, date)) or not isinstance(e, (int, float)):
            continue
        h = ws.cell(row=r, column=8).value
        change = (e - prev_e) if prev_e is not None else None
        rows.append({
            "date": d.strftime("%Y-%m-%d"),
            "share": round(e, 2),
            "change": round(change, 2) if change is not None else None,
            "mom": round(change / prev_e, 8) if (change is not None and prev_e) else None,
            "base": round(h, 2) if isinstance(h, (int, float)) else None,
            "rel_base": round(_rel_base(ws, r), 8) if r >= 3 else None,
        })
        prev_e = e
    return rows


def build_data() -> dict:
    """构建页面数据: summary(汇总) + details(各基金明细)。"""
    wb = openpyxl.load_workbook(TRACK_PATH, data_only=True)
    summary, details = [], {}
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        ws = wb[name]
        code = str(ws.cell(row=2, column=2).value or name).replace("\t", "").strip()
        fund_name = ws.cell(row=2, column=3).value or name
        detail = load_detail(ws)
        if not detail:
            continue
        details[name] = detail
        last = detail[-1]
        status, n = calc_status_from_detail(ws)
        cum = round(sum(_rel_base(ws, r) for r in range(3, ws.max_row + 1)[-n:]), 8) if n else None
        summary.append({
            "code": code,
            "name": str(fund_name).strip(),
            "date": last["date"],
            "share": last["share"],
            "change": last["change"],
            "mom": last["mom"],
            "rel_base": last["rel_base"],
            "status": status,
            "cum": cum,
        })
    return {"generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "summary": summary, "details": details}


# ---------- HTML 模板 ----------

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ETF 公示份额跟踪</title>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: "Microsoft YaHei", "PingFang SC", sans-serif; background: #f5f6f8; color: #2c2c2a; padding: 24px; }
  .wrap { max-width: 1100px; margin: 0 auto; }
  h1 { font-size: 20px; font-weight: 500; margin-bottom: 4px; }
  .sub { color: #888; font-size: 12px; margin-bottom: 20px; }
  .card { background: #fff; border-radius: 12px; padding: 20px; margin-bottom: 24px;
          box-shadow: 0 1px 3px rgba(0,0,0,.06); }
  .card h2 { font-size: 15px; font-weight: 500; margin-bottom: 14px; }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  th { background: #4a4a4a; color: #fff; font-weight: 500; padding: 9px 10px; text-align: right; white-space: nowrap; }
  th:first-child, th:nth-child(2), th:nth-child(8) { text-align: left; }
  td { padding: 8px 10px; border-bottom: 1px solid #f0f0f0; text-align: right; white-space: nowrap; }
  td:first-child, td:nth-child(2), td:nth-child(8) { text-align: left; }
  tbody tr:nth-child(even) { background: #fafafa; }
  .pos { color: #e24b4a; }   /* 涨 红 */
  .neg { color: #1d9e75; }   /* 跌 绿 */
  .zero { color: #999; }
  .pill { display: inline-block; padding: 2px 10px; border-radius: 10px; font-size: 12px; }
  .pill-up { background: #fcebeb; color: #a32d2d; }
  .pill-down { background: #e1f5ee; color: #0f6e56; }
  .pill-none { background: #f1efe8; color: #888; }
  .toolbar { display: flex; align-items: center; gap: 12px; margin-bottom: 14px; flex-wrap: wrap; }
  .toolbar label { font-size: 13px; color: #555; }
  select { padding: 6px 10px; border: 1px solid #ddd; border-radius: 8px; font-size: 13px;
           background: #fff; min-width: 220px; }
  .pager { display: flex; align-items: center; gap: 10px; margin-top: 14px; justify-content: flex-end;
           font-size: 13px; color: #555; }
  .pager button { padding: 5px 12px; border: 1px solid #ddd; border-radius: 8px; background: #fff;
                  cursor: pointer; font-size: 13px; }
  .pager button:hover { background: #f5f6f8; }
  .pager button:disabled { opacity: .4; cursor: not-allowed; }
  .muted { color: #999; }
</style>
</head>
<body>
<div class="wrap">
  <h1>ETF 公示份额跟踪</h1>
  <div class="sub">数据生成时间：__GENERATED__　·　数据来源：上交所 / 深交所　·　单位：万份</div>

  <div class="card">
    <h2>汇总表</h2>
    <table id="summaryTable">
      <thead><tr>
        <th>代码</th><th>名称</th><th>最新日期</th><th>公示份额</th><th>当日变化</th>
        <th>环比昨日变动</th><th>相对基期变动</th><th>状态</th><th>连续n日累计变动</th>
      </tr></thead>
      <tbody></tbody>
    </table>
  </div>

  <div class="card">
    <h2>明细表</h2>
    <div class="toolbar">
      <label for="fundSelect">基金：</label>
      <select id="fundSelect"></select>
      <span class="muted" id="detailHint"></span>
    </div>
    <table id="detailTable">
      <thead><tr>
        <th>日期</th><th>当日公示份额</th><th>当日变化</th><th>环比变动比例</th>
        <th>基期总份额</th><th>相对基期变动比例</th>
      </tr></thead>
      <tbody></tbody>
    </table>
    <div class="pager" id="pager"></div>
  </div>
</div>

<script>
window.DATA = __DATA__;
</script>
<script>
(function () {
  var DATA = window.DATA;
  var PAGE_SIZE = 10;
  var summary = DATA.summary || [];
  var details = DATA.details || {};
  var currentFund = null;
  var currentPage = 1;

  function fmtNum(v) {
    if (v === null || v === undefined) return '-';
    var neg = v < 0;
    var s = Math.abs(v).toFixed(2);
    s = s.replace(/\\B(?=(\\d{3})+(?!\\d))/g, ',');
    return (neg ? '-' : '') + s;
  }
  function fmtPct(v) {
    if (v === null || v === undefined) return '-';
    return (v * 100).toFixed(2) + '%';
  }
  function cls(v) {
    if (v === null || v === undefined) return 'zero';
    if (v > 0) return 'pos';
    if (v < 0) return 'neg';
    return 'zero';
  }
  function numCell(v) {
    return '<td class="' + cls(v) + '">' + fmtNum(v) + '</td>';
  }
  function pctCell(v) {
    return '<td class="' + cls(v) + '">' + fmtPct(v) + '</td>';
  }
  function pill(status) {
    var s = status || '-';
    var c = 'pill-none';
    if (s.indexOf('增持') > -1) c = 'pill-up';
    else if (s.indexOf('减持') > -1) c = 'pill-down';
    return '<span class="pill ' + c + '">' + s + '</span>';
  }

  function renderSummary() {
    var tb = document.querySelector('#summaryTable tbody');
    var html = '';
    summary.forEach(function (r) {
      html += '<tr>'
        + '<td>' + r.code + '</td>'
        + '<td>' + r.name + '</td>'
        + '<td>' + (r.date || '-') + '</td>'
        + '<td>' + fmtNum(r.share) + '</td>'
        + numCell(r.change)
        + pctCell(r.mom)
        + pctCell(r.rel_base)
        + '<td>' + pill(r.status) + '</td>'
        + pctCell(r.cum)
        + '</tr>';
    });
    tb.innerHTML = html;
  }

  function renderSelect() {
    var sel = document.querySelector('#fundSelect');
    sel.innerHTML = '';
    summary.forEach(function (r, i) {
      var opt = document.createElement('option');
      opt.value = r.code;
      opt.textContent = r.name + '（' + r.code + '）';
      sel.appendChild(opt);
    });
    sel.value = summary.length ? summary[0].code : '';
    sel.addEventListener('change', function () {
      currentFund = sel.value;
      currentPage = 1;
      renderDetail();
    });
  }

  function renderDetail() {
    var rows = (details[currentFund] || []).slice().reverse(); /* 新→旧 */
    var tb = document.querySelector('#detailTable tbody');
    var total = rows.length;
    var pages = Math.max(1, Math.ceil(total / PAGE_SIZE));
    if (currentPage > pages) currentPage = pages;
    var start = (currentPage - 1) * PAGE_SIZE;
    var slice = rows.slice(start, start + PAGE_SIZE);
    var html = '';
    slice.forEach(function (r) {
      html += '<tr>'
        + '<td>' + r.date + '</td>'
        + '<td>' + fmtNum(r.share) + '</td>'
        + numCell(r.change)
        + pctCell(r.mom)
        + '<td>' + fmtNum(r.base) + '</td>'
        + pctCell(r.rel_base)
        + '</tr>';
    });
    tb.innerHTML = html || '<tr><td colspan="6" class="muted">暂无数据</td></tr>';
    document.querySelector('#detailHint').textContent = '共 ' + total + ' 条记录';
    renderPager(pages);
  }

  function renderPager(pages) {
    var p = document.querySelector('#pager');
    p.innerHTML = '<button id="prevBtn">上一页</button>'
      + '<span>' + currentPage + ' / ' + pages + '</span>'
      + '<button id="nextBtn">下一页</button>';
    var prev = document.querySelector('#prevBtn');
    var next = document.querySelector('#nextBtn');
    prev.disabled = currentPage <= 1;
    next.disabled = currentPage >= pages;
    prev.onclick = function () { if (currentPage > 1) { currentPage--; renderDetail(); } };
    next.onclick = function () { if (currentPage < pages) { currentPage++; renderDetail(); } };
  }

  function init() {
    renderSummary();
    renderSelect();
    if (summary.length) { currentFund = summary[0].code; renderDetail(); }
    else {
      document.querySelector('#detailTable tbody').innerHTML =
        '<tr><td colspan="6" class="muted">暂无基金数据</td></tr>';
    }
  }
  init();
})();
</script>
</body>
</html>
"""


def render_html(data: dict) -> str:
    payload = json.dumps(data, ensure_ascii=False)
    payload = payload.replace("</", "<\\/")      # 防止 </script> 中断
    return HTML_TEMPLATE.replace("__DATA__", payload) \
                        .replace("__GENERATED__", data["generated_at"])


def main():
    data = build_data()
    os.makedirs(os.path.dirname(OUT_HTML), exist_ok=True)
    with open(OUT_HTML, "w", encoding="utf-8") as f:
        f.write(render_html(data))
    # 详细日志：方便 CI/本地排查"页面数据是否更新、更新到哪天"
    log.info("已生成 %s (generated_at=%s, %d 只基金)",
             OUT_HTML, data["generated_at"], len(data["summary"]))
    for s in data["summary"]:
        rows = data["details"].get(s["code"], [])
        last_date = rows[-1]["date"] if rows else "-"
        log.info("  [%s] %s 明细行数=%d 最新日期=%s 份额=%s",
                 s["code"], s["name"], len(rows), last_date, s["share"])


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log.error("导出失败: %s", e)
        raise
