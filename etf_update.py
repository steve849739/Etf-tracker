#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ETF 跟踪 Excel 写入层——读抓取 JSON，写明细 sheet + 计算公式 + 总结。

依赖 etf_fetch.py 输出的 data/etf_data.json。

用法:
  python etf_update.py                          -> daily 模式: 追加最新数据 + 刷新总结
  python etf_update.py --backfill 起 止            -> 重建明细 + 总结(JSON中history字段)

环境变量: ETF_TRACK_PATH(跟踪表路径, 默认 etf跟踪.xlsx) ETF_DATA_PATH(JSON路径)
"""
import sys
import os
import json
import logging
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

TRACK_PATH = os.environ.get("ETF_TRACK_PATH", os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "etf跟踪.xlsx"))
DATA_PATH = os.environ.get("ETF_DATA_PATH", os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "data", "etf_data.json"))
SUMMARY_SHEET = "总结"

COL_HEADERS = ["来源", "代码", "名称", "日期", "当日公示份额（万份）",
               "当日变化", "环比变动比例", "基期总份额", "相对基期变动比例"]
SUMMARY_HEADERS = ["代码", "名称", "最新日期", "公示份额", "当日变化",
                   "环比昨日变动", "相对基期变动", "状态", "连续n日累计变动比例"]

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("etf-update")


# ---------- 明细 sheet ----------

def ensure_sheet(wb, code: str):
    if code in wb.sheetnames:
        return wb[code]
    ws = wb.create_sheet(code)
    for c, h in enumerate(COL_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    return ws


def update_sheet(wb, fund: dict, stat_date: str, share: float) -> str:
    """追加一行到明细 sheet, 写入公式。返回结果描述。"""
    code = fund["code"]
    ws = ensure_sheet(wb, code)
    last_row = 1
    for r in range(ws.max_row, 0, -1):
        if ws.cell(row=r, column=1).value is not None:
            last_row = r; break
    d = datetime.strptime(stat_date, "%Y-%m-%d")
    for r in range(2, last_row + 1):
        dv = ws.cell(row=r, column=4).value
        if isinstance(dv, (datetime, date)) and dv.date() == d.date():
            return "[%s] %s 已存在(第%d行)" % (code, stat_date, r)
    b_prefix = ""
    if last_row >= 2:
        bv = ws.cell(row=2, column=2).value
        if isinstance(bv, str) and bv.startswith("\t"):
            b_prefix = "\t"
    new_row = last_row + 1
    ws.cell(row=new_row, column=1, value=fund["source"] if fund.get("source") else "上交所" if code.startswith("51") else "深交所")
    ws.cell(row=new_row, column=2, value=b_prefix + code)
    ws.cell(row=new_row, column=3, value=fund["name"])
    ws.cell(row=new_row, column=4, value=d)
    ws.cell(row=new_row, column=5, value=share)
    if last_row >= 2:
        ws.cell(row=new_row, column=6, value="=E%d-E%d" % (new_row, last_row))
        ws.cell(row=new_row, column=7, value="=F%d/E%d" % (new_row, last_row))
        base = ws.cell(row=last_row, column=8).value
        if base is not None:
            ws.cell(row=new_row, column=8, value=base)
            ws.cell(row=new_row, column=9, value="=F%d/H%d" % (new_row, new_row))
    return "[%s] 已追加 %s 份额 %.2f 万份 -> 第%d行" % (code, stat_date, share, new_row)


def rebuild_sheet(wb, fund: dict, records: dict, base, period_desc: str) -> str:
    """重建明细: 表头 + records(升序) + 公式。"""
    code = fund["code"]
    if code in wb.sheetnames:
        del wb[code]
    ws = wb.create_sheet(code)
    for c, h in enumerate(COL_HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="B50005")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sorted_dates = sorted(records.keys())
    for i, ds in enumerate(sorted_dates):
        r = i + 2
        ws.cell(row=r, column=1, value=fund["source"] if fund.get("source") else "上交所" if code.startswith("51") else "深交所")
        ws.cell(row=r, column=2, value=code)
        ws.cell(row=r, column=3, value=fund["name"])
        ws.cell(row=r, column=4, value=datetime.strptime(ds, "%Y-%m-%d"))
        ws.cell(row=r, column=5, value=records[ds])
        if i >= 1:
            ws.cell(row=r, column=6, value="=E%d-E%d" % (r, r - 1))
            ws.cell(row=r, column=7, value="=F%d/E%d" % (r, r - 1))
        if base is not None:
            ws.cell(row=r, column=8, value=base)
            if i >= 1:
                ws.cell(row=r, column=9, value="=F%d/H%d" % (r, r))
    return "[%s] 重建完成: %d 交易日(%s)%s" % (
        code, len(sorted_dates), period_desc,
        ", 基期%.2f" % base if base is not None else ", 无基期")


# ---------- 总结 ----------

def calc_status_from_detail(ws) -> tuple:
    """返回 (状态文本, n)。"""
    vals = []
    for r in range(3, ws.max_row + 1):
        ep = ws.cell(row=r - 1, column=5).value
        e = ws.cell(row=r, column=5).value
        h = ws.cell(row=r, column=8).value
        if not all(isinstance(v, (int, float)) for v in (ep, e, h)) or h == 0:
            continue
        vals.append((r, (e - ep) / h))
    if len(vals) < 2:
        return "-", None
    cur, prev = vals[-1][1], vals[-2][1]
    if cur >= 0:
        if prev < 0:
            return "反转增持", 1
        x = 0
        for _, v in reversed(vals):
            if v >= 0: x += 1
            else: break
        return "连续%d日增持" % x, x
    else:
        if prev >= 0:
            return "反转减持", 1
        x = 0
        for _, v in reversed(vals):
            if v < 0: x += 1
            else: break
        return "连续%d日减持" % x, x


def _sheet_ref(name: str) -> str:
    return "'%s'" % name if not name[0].isalpha() else name


def update_summary(wb, funds: list) -> list:
    """更新总结 sheet, 返回使用行号列表。"""
    if SUMMARY_SHEET in wb.sheetnames:
        ws = wb[SUMMARY_SHEET]
    else:
        ws = wb.create_sheet(SUMMARY_SHEET)
    if ws.max_row == 1:
        for c, h in enumerate(SUMMARY_HEADERS, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="B50005")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    row_map = {}
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a: row_map[str(a).strip()] = r
    used = []
    for fund in funds:
        code = fund["code"]
        if code not in wb.sheetnames:
            continue
        ref = _sheet_ref(code)
        r = row_map.get(code)
        if r is None:
            r = ws.max_row + 1
        used.append(r)
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=fund["name"])
        ws.cell(row=r, column=3, value="=INDEX(%s!$D:$D,COUNTA(%s!$D:$D))" % (ref, ref))
        ws.cell(row=r, column=4, value="=INDEX(%s!$E:$E,COUNTA(%s!$E:$E))" % (ref, ref))
        ws.cell(row=r, column=5, value="=INDEX(%s!$F:$F,COUNTA(%s!$E:$E))" % (ref, ref))
        ws.cell(row=r, column=6, value="=INDEX(%s!$G:$G,COUNTA(%s!$E:$E))" % (ref, ref))
        ws.cell(row=r, column=7, value="=INDEX(%s!$I:$I,COUNTA(%s!$E:$E))" % (ref, ref))
        ws.cell(row=r, column=8, value="")
        status, n = calc_status_from_detail(wb[code])
        ws.cell(row=r, column=8, value=status)
        if n:
            ws.cell(row=r, column=9,
                    value="=SUM(INDEX(%s!$I:$I,COUNTA(%s!$E:$E)-%d+1):INDEX(%s!$I:$I,COUNTA(%s!$E:$E)))"
                          % (ref, ref, n, ref, ref))
        else:
            ws.cell(row=r, column=9, value=None)
        ws.cell(row=r, column=3).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=4).number_format = "#,##0.00"
        ws.cell(row=r, column=5).number_format = "#,##0.00"
        ws.cell(row=r, column=6).number_format = "0.00%"
        ws.cell(row=r, column=7).number_format = "0.00%"
        ws.cell(row=r, column=9).number_format = "0.00%"
    return used


# ---------- main ----------

def main():
    if not os.path.exists(DATA_PATH):
        log.error("数据文件不存在: %s (请先运行 etf_fetch.py)", DATA_PATH)
        sys.exit(1)
    with open(DATA_PATH, "r", encoding="utf-8") as f:
        payload = json.load(f)
    funds_in = payload.get("funds", [])
    mode = payload.get("mode", "daily")
    backfill = "--backfill" in sys.argv
    start, end = None, None
    if backfill:
        idx = sys.argv.index("--backfill")
        start, end = sys.argv[idx + 1], sys.argv[idx + 2]

    wb = openpyxl.load_workbook(TRACK_PATH)
    results = []

    for item in funds_in:
        code, name, source = item["code"], item["name"], item.get("source", "")
        # 补 source 信息
        fund_ctx = {"name": name, "code": code, "source": source}
        base = item.get("base_share")
        try:
            if mode == "daily":
                latest = item.get("latest")
                if not latest or latest.get("share") is None:
                    results.append("[%s] 无数据(抓取失败或跳过)" % code)
                    continue
                sd, sh = latest["date"], latest["share"]
                log.info("[%s] %s=%.2f万份", code, sd, sh)
                results.append(update_sheet(wb, fund_ctx, sd, sh))
            elif mode == "backfill":
                hist = item.get("history", {})
                if not hist:
                    results.append("[%s] 无历史数据" % code)
                    continue
                log.info("[%s] %s~%s: %d天", code, start, end, len(hist))
                results.append(rebuild_sheet(wb, fund_ctx, hist, base,
                                             "%s~%s" % (start, end)))
        except Exception as e:
            log.error("[%s] 失败: %s", code, e)
            results.append("[%s] 失败: %s" % (code, e))

    update_summary(wb, funds_in)
    wb.save(TRACK_PATH)
    log.info("保存完成: %s", TRACK_PATH)
    for line in results:
        log.info(line)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log.error("更新失败: %s", e)
        sys.exit(1)
